"""
RiskSync CRQ Monte Carlo Engine
================================
Reads Frequency and Severity parameters directly from the
RiskSync_CRQ_DataWarehouse workbook and runs a
Poisson(frequency) x Lognormal(severity) Monte Carlo simulation,
following the exact method defined in the workbook's Assumptions sheet:

    Frequency_Distribution : Poisson
    Severity_Distribution  : Lognormal
    Severity_Sigma_Method  : sigma = ln(P95 / Median) / 1.645
    Simulation_Runs        : 100,000
    Random_Seed            : 42

FIXED in this version (v7):
  - The workbook does NOT contain "=== TABLE A ===" style text markers.
    Table boundaries are actually delimited by BLANK rows, with the section
    header repeating a column-1 label ("Industry_ID" appears twice in
    Frequency_IRIS -- once for Table A, once for Table B -- then
    "Revenue_Class" for Table C). The previous version's marker-based
    parser silently produced empty dicts for baseline/ind_mult/rev_mult/
    ind_sev/rev_sev, which crashed get_frequency()/get_severity_params()
    with a TypeError/ValueError as soon as main() ran. This version parses
    by (a) blank-row boundaries and (b) which known header repeats, which
    matches the actual workbook layout confirmed by direct inspection.
  - WORKBOOK_PATH corrected to resolve relative to this script's own
    directory (previously a hardcoded /mnt/user-data/outputs/ path that
    did not match the actual delivery folder).

Usage:
    python crq_monte_carlo.py --industry Financial --revenue "$1B to $10B"
    python crq_monte_carlo.py --industry Financial --revenue "$1B to $10B" --ssc_grade F

Notes / current data limitations (see chat for full discussion):
  - Frequency baseline (absolute annual probability) is only published
    for 8 of 20 IRIS industries (+ 1 "Overall" row). For the other 12,
    this script falls back to Overall baseline (9.3%) x Industry_Multiplier
    as an APPROXIMATION. This is flagged in the output (approx=True).
  - Industry_Master exists and is confidence-labeled, but this script still
    matches industry names directly against Severity_IRIS / Frequency_IRIS
    (IRIS naming only) for the core calculation. DBIR / IBM / NetDiligence
    are not yet blended into the severity distribution itself (still pending
    - Primary/Secondary double-counting issue flagged earlier in the chat).
  - Frequency_SSC is based on a SINGLE source (one SecurityScorecard slide),
    not independently cross-validated. Treat SSC-adjusted results as
    directional, not precise.
"""

import argparse
import math
import os
import numpy as np
import openpyxl

# Resolve relative to this script's own folder, not the caller's cwd.
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
WORKBOOK_PATH = os.path.join(_THIS_DIR, "RiskSync_CRQ_DataWarehouse_LATEST.xlsx")

Z_95 = 1.645  # z-score for the 95th percentile, per Assumptions sheet


def _iter_blank_delimited_sections(rows):
    """
    Yield successive (header_row, data_rows) blocks from a sheet's rows,
    where blocks are separated by fully-blank rows (col A is None) and/or
    a leading fully-blank row at the very top. Skips leading/trailing
    blank rows. This matches the actual workbook layout: no text markers,
    just blank-row-delimited tables each starting with its own header row.
    """
    current = []
    for row in rows:
        is_blank = row[0] is None and all(v is None for v in row)
        if is_blank:
            if current:
                yield current[0], current[1:]
                current = []
            continue
        current.append(row)
    if current:
        yield current[0], current[1:]


def load_frequency_tables(wb):
    ws = wb["Frequency_IRIS"]
    rows = list(ws.iter_rows(values_only=True))

    baseline = {}   # Table A: Industry -> absolute annual frequency (or None)
    ind_mult = {}   # Table B: Industry -> relative multiplier
    rev_mult = {}   # Table C: Revenue class -> multiplier

    blocks = list(_iter_blank_delimited_sections(rows))
    seen_industry_id_headers = 0
    for header, data_rows in blocks:
        label0 = header[0]
        if label0 == "Industry_ID":
            seen_industry_id_headers += 1
            if seen_industry_id_headers == 1:
                for r in data_rows:
                    ind, freq = r[0], r[1]
                    if ind is None:
                        continue
                    baseline[ind] = None if freq in (None, "N/A") else float(freq)
            else:
                for r in data_rows:
                    ind, mult = r[0], r[1]
                    if ind is None:
                        continue
                    ind_mult[ind] = float(mult)
        elif label0 == "Revenue_Class":
            for r in data_rows:
                rc, mult = r[0], r[1]
                if rc is None:
                    continue
                rev_mult[rc] = float(mult)

    return baseline, ind_mult, rev_mult


def load_severity_tables(wb):
    ws = wb["Severity_IRIS"]
    rows = list(ws.iter_rows(values_only=True))

    ind_sev = {}   # Industry -> (median, p95)
    rev_sev = {}   # Revenue class -> (typical, extreme)  [used as median/P95 proxy]

    blocks = list(_iter_blank_delimited_sections(rows))
    for header, data_rows in blocks:
        label0 = header[0]
        if label0 == "Industry_ID":
            for r in data_rows:
                ind = r[0]
                if ind is None:
                    continue
                median, p95 = r[2], r[3]
                ind_sev[ind] = (float(median), float(p95))
        elif label0 == "Revenue_Class":
            for r in data_rows:
                rc = r[0]
                if rc is None:
                    continue
                typical, extreme = r[1], r[2]
                rev_sev[rc] = (float(typical), float(extreme))

    return ind_sev, rev_sev


SCENARIO_COVERAGE = {
    "ransomware": True,               # A.5 Cyber-extortion
    "bec": True,                      # A.6 Cyber-crime
    "data_breach": True,              # A.1 Incident and breach response
    "business_interruption": True,    # A.3 (subject to 12h waiting period, not modeled as $ deduction)
    "third_party_liability": True,    # B.1/B.2
    "state_sponsored_attack": False,  # War/Cyber War/Cyber Operation Exclusion (LMA5566)
    "cryptocurrency_loss": False,     # Exclusion #17
    "intentional_act_by_insured": False,  # Exclusion #6
}


def is_covered(scenario_type):
    if scenario_type not in SCENARIO_COVERAGE:
        raise ValueError(f"Unknown scenario_type '{scenario_type}'. "
                          f"Available: {sorted(SCENARIO_COVERAGE)}")
    return SCENARIO_COVERAGE[scenario_type]


def load_ssc_multipliers(wb):
    ws = wb["Frequency_SSC"]
    rows = list(ws.iter_rows(values_only=True))
    mult = {}
    for row in rows[1:]:
        grade = row[0]
        if grade is None:
            continue
        mult[grade] = float(row[1])
    return mult


def lognormal_params(median, p95):
    sigma = math.log(p95 / median) / Z_95
    mu = math.log(median)
    return mu, sigma


def get_frequency(industry, revenue_class, baseline, ind_mult, rev_mult,
                   ssc_mult_table=None, ssc_grade=None):
    overall = baseline.get("Overall (All Industries)")
    if overall is None:
        raise ValueError("Could not find 'Overall (All Industries)' baseline in Frequency_IRIS Table A.")

    approx = False
    if industry not in baseline:
        raise ValueError(f"Unknown industry '{industry}'. Available: {sorted(baseline)}")

    if baseline[industry] is not None:
        base = baseline[industry]
    else:
        base = overall * ind_mult.get(industry, 1.0)
        approx = True

    if revenue_class not in rev_mult:
        raise ValueError(f"Unknown revenue_class '{revenue_class}'. Available: {sorted(rev_mult)}")
    rev_m = rev_mult[revenue_class]
    lef = base * rev_m

    ssc_m = 1.0
    if ssc_grade is not None:
        if ssc_mult_table is None or ssc_grade not in ssc_mult_table:
            raise ValueError(f"Unknown SSC grade '{ssc_grade}'. "
                              f"Available: {sorted(ssc_mult_table) if ssc_mult_table else '[]'}")
        ssc_m = ssc_mult_table[ssc_grade]
        lef *= ssc_m

    return lef, approx, ssc_m


def get_severity_params(industry, revenue_class, ind_sev, rev_sev):
    if industry not in ind_sev:
        raise ValueError(f"No severity data for industry '{industry}'. "
                          f"Available: {sorted(ind_sev)}")
    median, p95 = ind_sev[industry]
    mu, sigma = lognormal_params(median, p95)

    rev_note = None
    if revenue_class in rev_sev:
        rev_median, rev_p95 = rev_sev[revenue_class]
        rev_note = (rev_median, rev_p95)

    return mu, sigma, rev_note


def run_monte_carlo(lef, mu, sigma, n_runs=100_000, seed=42):
    rng = np.random.default_rng(seed)
    n_events = rng.poisson(lam=lef, size=n_runs)
    total_losses = np.zeros(n_runs)

    max_events = n_events.max()
    if max_events > 0:
        all_sev = rng.lognormal(mean=mu, sigma=sigma, size=(n_runs, max_events))
        mask = np.arange(max_events) < n_events[:, None]
        total_losses = (all_sev * mask).sum(axis=1)

    return total_losses


def summarize(losses):
    pct = lambda p: np.percentile(losses, p)
    return {
        "mean_ALE": losses.mean(),
        "median": pct(50),
        "p75": pct(75),
        "p90": pct(90),
        "p95": pct(95),
        "p99": pct(99),
        "max_simulated": losses.max(),
        "prob_zero_loss_year": (losses == 0).mean(),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--industry", required=True, help="e.g. Financial, Healthcare, Retail")
    parser.add_argument("--revenue", required=True,
                         help='e.g. "$1B to $10B" (must match Frequency_IRIS Table C labels)')
    parser.add_argument("--ssc_grade", default=None, choices=["A", "B", "C", "D", "F"],
                         help="Optional SecurityScorecard grade to adjust LEF (Vulnerability proxy)")
    parser.add_argument("--scenario", default=None, choices=list(SCENARIO_COVERAGE.keys()),
                         help="Optional scenario type to check Yes/No policy coverage (no $ adjustment)")
    parser.add_argument("--runs", type=int, default=100_000)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    baseline, ind_mult, rev_mult = load_frequency_tables(wb)
    ind_sev, rev_sev = load_severity_tables(wb)
    ssc_mult_table = load_ssc_multipliers(wb)

    lef, approx, ssc_m = get_frequency(args.industry, args.revenue, baseline, ind_mult, rev_mult,
                                        ssc_mult_table=ssc_mult_table, ssc_grade=args.ssc_grade)
    mu, sigma, rev_note = get_severity_params(args.industry, args.revenue, ind_sev, rev_sev)

    losses = run_monte_carlo(lef, mu, sigma, n_runs=args.runs, seed=args.seed)
    stats = summarize(losses)

    print("=" * 60)
    print(f"CRQ Monte Carlo Result — {args.industry} / {args.revenue}"
          f"{' / SSC=' + args.ssc_grade if args.ssc_grade else ''}")
    print("=" * 60)
    print(f"Annual Loss Event Frequency (LEF): {lef:.4f}  "
          f"{'[APPROXIMATED via Overall x Industry multiplier]' if approx else '[from published IRIS baseline]'}")
    if args.ssc_grade:
        print(f"  SSC adjustment applied: grade={args.ssc_grade}, multiplier={ssc_m}x  "
              f"[single-source, directional only]")
    if args.scenario:
        covered = is_covered(args.scenario)
        print(f"Policy coverage check ({args.scenario}): "
              f"{'COVERED' if covered else 'EXCLUDED'}  "
              f"[Yes/No only, per Policy_Coverage sheet - no dollar amounts modeled]")
    print(f"Severity lognormal params: mu={mu:.3f}, sigma={sigma:.3f}")
    if rev_note:
        print(f"  (Revenue-tier Typical/Extreme loss for cross-check: "
              f"${rev_note[0]:,.0f} / ${rev_note[1]:,.0f})")
    print("-" * 60)
    for k, v in stats.items():
        if "prob" in k:
            print(f"{k:25s}: {v:.2%}")
        else:
            print(f"{k:25s}: ${v:,.0f}")
    print("=" * 60)

    return losses, stats


if __name__ == "__main__":
    main()
