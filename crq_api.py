"""
RiskSync CRQ Calculation API (backend only — no frontend in this step)
======================================================================
Thin JSON layer over the existing `crq_monte_carlo.py` engine. Reuses the
engine's own parsing/calculation functions (does NOT re-derive the model).

Implements PROJECT_BRIEF.md "Interaction flow (Step 0)" input validation:
  - Industry  : REQUIRED. Accepts either the Industry_Master display name
                (e.g. "Financial Services") or an IRIS name (e.g. "Financial").
                Resolution uses ONLY the Industry_Master sheet's own columns —
                IRIS_Frequency_Industry (for Frequency_IRIS lookups) and
                IRIS_Severity_Industry (for Severity_IRIS lookups); IRIS 2025
                itself labels one industry differently across its frequency
                and severity figures ("Other" vs "Other services"), so the
                per-sheet name lives in the workbook, not in code. Orphan
                industries (no IRIS anchor) are rejected with an error.
  - Revenue   : REQUIRED. Must match a Frequency_IRIS Table C label exactly.
  - SSC grade : OPTIONAL. If given, must exist in the Frequency_SSC sheet
                (A/B/C/D/F). Default = no SSC adjustment applied.
Monte Carlo is never run on partial input — validation errors are returned
as JSON with an "error" field instead.

Coverage: Yes/No only, from the engine's SCENARIO_COVERAGE dict (per the
Policy_Coverage sheet). No dollar amounts are modeled — the specimen policy
has none (see PROJECT_BRIEF "What NOT to do").

Model Confidence rule (documented per PROJECT_BRIEF Layer 1, which asks to
derive it from Industry_Master Mapping_Confidence + Evidence_Register SQR
of the sources actually used):
  High   = published IRIS baseline (approx=False) AND Mapping_Confidence High
  Medium = published baseline with Medium mapping, OR approximated baseline
           with High mapping
  Low    = everything else (mapping Low/None, or approximated + mapping <= Medium)
  If an SSC grade is applied, the level is downgraded one step, because
  Frequency_SSC's source (EV_05) is SQR=Low / single-source.

Usage:
    python crq_api.py --industry "Financial Services" --revenue "$1B to $10B"
    python crq_api.py --industry Education --revenue "$10M to $100M"
    python crq_api.py --industry Financial --revenue "$1B to $10B" --ssc_grade F --compact
Output: JSON on stdout (exit code 1 on validation error).
"""

import argparse
import json
import math
import sys

import numpy as np
import openpyxl

from crq_monte_carlo import (
    SCENARIO_COVERAGE,
    WORKBOOK_PATH,
    Z_95,
    get_frequency,
    get_severity_params,
    load_frequency_tables,
    load_severity_tables,
    load_ssc_multipliers,
    run_monte_carlo,
    summarize,
)


class ValidationError(ValueError):
    """Raised for Step 0 input-validation failures (returned as JSON errors)."""


def load_industry_master(wb):
    ws = wb["Industry_Master"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}
    industries = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        industries.append({
            "industry_id": r[idx["Industry_ID"]],
            "display_name": r[idx["RiskSync_Industry"]],
            "iris_frequency_name": r[idx["IRIS_Frequency_Industry"]],
            "iris_severity_name": r[idx["IRIS_Severity_Industry"]],
            "dbir_industry": r[idx["DBIR_Industry"]],
            "mapping_confidence": r[idx["Mapping_Confidence"]],
        })
    return industries


def load_revenue_severity_t1(wb):
    """Severity_Revenue_T1 sheet: IRIS Table 1 transcription + tier mapping.

    Returns (t1_table, tier_map):
      t1_table: T1 tier -> (median, p95)
      tier_map: Frequency_IRIS Table C label -> T1 tier
    """
    ws = wb["Severity_Revenue_T1"]
    rows = list(ws.iter_rows(values_only=True))
    t1_table, tier_map = {}, {}
    for header, data_rows in _iter_blank_delimited_sections_api(rows):
        if header[0] == "T1_Revenue_Tier":
            for r in data_rows:
                if r[0] is not None:
                    t1_table[r[0]] = (float(r[1]), float(r[2]))
        elif header[0] == "Revenue_Class":
            for r in data_rows:
                if r[0] is not None:
                    tier_map[r[0]] = r[1]
    return t1_table, tier_map


def _iter_blank_delimited_sections_api(rows):
    current = []
    for row in rows:
        if all(v is None for v in row):
            if current:
                yield current[0], current[1:]
                current = []
            continue
        current.append(row)
    if current:
        yield current[0], current[1:]


def revenue_severity_multiplier(revenue_class, t1_table, tier_map):
    """Design A (SEVERITY_ADJUSTMENT_DESIGN.md): M = T1 median / GM(4 medians).

    GM normalization = sum-to-zero constraint in log space, so the four
    multipliers are exposure-neutral on average (their product is 1).
    """
    tier = tier_map[revenue_class]
    medians = [m for m, _ in t1_table.values()]
    gm = math.exp(sum(math.log(m) for m in medians) / len(medians))
    m_rev = t1_table[tier][0] / gm
    return m_rev, tier, gm


def load_cost_component_breakdown(wb):
    """Severity_NetDiligence: the "Table B: Cost Component Breakdown" block,
    located by its title text (not row/table position — the sheet may grow
    additional sections), per PROJECT_BRIEF's Layer 2 Loss Breakdown spec.

    Returns a list of component dicts (Component, Min/Avg/Max, Currency,
    Primary_Secondary, Basis, Source_Report, Source_Section, Notes,
    Applicable_Industry). Applicable_Industry is a workbook column (not a
    hardcoded string in code) — see Industry_Master precedent for why
    per-source industry applicability lives in the data, not in Python.
    """
    ws = wb["Severity_NetDiligence"]
    rows = list(ws.iter_rows(values_only=True))
    title_idx = None
    for i, row in enumerate(rows):
        if row[0] and str(row[0]).startswith("Table B: Cost Component Breakdown"):
            title_idx = i
            break
    if title_idx is None:
        raise ValidationError(
            "Severity_NetDiligence: 'Table B: Cost Component Breakdown' "
            "section not found (workbook layout drift).")
    header = rows[title_idx + 1]
    idx = {name: i for i, name in enumerate(header) if name is not None}
    required = ("Component", "Avg", "Primary_Secondary", "Applicable_Industry")
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValidationError(
            f"Severity_NetDiligence Cost Component Breakdown is missing "
            f"expected column(s): {missing}.")
    components = []
    for row in rows[title_idx + 2:]:
        if row[0] is None:
            break
        components.append({k: row[i] for k, i in idx.items()})
    return components


def load_cause_of_loss(wb):
    """Severity_NetDiligence Cause_of_Loss (SME) table, extended with
    Total/Pct_of_Total_Cost columns (Table 9). All-SME aggregate across every
    industry (NOT industry-specific) — used for the Layer 2 "Top Causes of
    Loss" reference list, replacing the old illustrative/generic risk-driver
    examples with actual claims-study statistics.
    """
    ws = wb["Severity_NetDiligence"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] == "Cause_of_Loss":
            header_idx = i
            break
    if header_idx is None:
        raise ValidationError("Severity_NetDiligence: 'Cause_of_Loss' table not found.")
    header = rows[header_idx]
    idx = {name: i for i, name in enumerate(header) if name is not None}
    causes = []
    for row in rows[header_idx + 1:]:
        if row[0] is None:
            break
        if row[idx.get("Pct_of_Total_Cost", -1)] is None:
            continue  # not yet extended with Table 9's Total/% columns
        causes.append({
            "cause": row[idx["Cause_of_Loss"]], "claims": row[idx["Claims"]],
            "avg_incident_cost_usd": row[idx["Avg_Incident_Cost"]],
            "total_incident_cost_usd": row[idx["Total_Incident_Cost"]],
            "pct_of_total_cost": row[idx["Pct_of_Total_Cost"]],
        })
    causes.sort(key=lambda c: c["pct_of_total_cost"], reverse=True)
    return causes


def load_incident_cost_context(wb):
    """Severity_NetDiligence: SME 2020-2024 all-industry aggregate reference
    figures for Business Interruption, Legal Costs, Recovery Expense, and
    Crisis Services' own share of total Incident Cost. These come from
    DIFFERENT claim subsets/tables than Table B's Crisis Services breakdown
    (e.g. the BI figure is the average for BI-affected claims specifically,
    not all claims) — so they are surfaced as independent reference stats,
    not merged into one shared "Incident Cost pie" with the Crisis Services
    donut, which would imply a common denominator the source data does not
    support.
    """
    ws = wb["Severity_NetDiligence"]
    rows = list(ws.iter_rows(values_only=True))
    result = {}
    for i, row in enumerate(rows):
        if row[0] == "Business Interruption Claims (SME)":
            result["bi_incident_cost_usd"] = row[1]
            result["bi_cost_usd"] = row[2]
            result["bi_crisis_cost_usd"] = row[3]
            result["bi_n_claims"] = row[4]
        elif row[0] == "Entity" and rows[i + 1][0] == "SME (< $2B revenue)":
            legal_row = rows[i + 1]
            result["legal_settlement_usd"] = legal_row[1]
            result["legal_defense_usd"] = legal_row[2]
            result["regulatory_fines_usd"] = legal_row[3]
            result["regulatory_defense_usd"] = legal_row[4]
            result["legal_total_usd"] = legal_row[5]  # Total_Legal_Cost
            result["legal_n_claims"] = legal_row[7]
        elif row[0] == "Avg_Recovery_Expense":
            result["recovery_expense_usd"] = row[1]
        elif row[0] == "Avg_Incident_Cost_Recovery_Context":
            result["recovery_incident_cost_usd"] = row[1]
        elif row[0] == "Crisis_Services_Pct_of_Incident_Cost":
            result["crisis_services_pct"] = row[1]
        elif row[0] == "Recovery_N_Claims":
            result["recovery_n_claims"] = row[1]
        elif row[0] == "Total_SME_Claims_Table5":
            result["total_sme_claims"] = row[1]
        elif row[0] == "BI_Yearly" and isinstance(row[1], int):
            result.setdefault("bi_yearly", []).append(
                {"year": row[1], "incident_cost_usd": row[2], "bi_cost_usd": row[3]})
        elif row[0] == "Recovery_Yearly" and isinstance(row[1], int):
            result.setdefault("recovery_yearly", []).append(
                {"year": row[1], "incident_cost_usd": row[2], "recovery_expense_usd": row[3]})
        elif row[0] == "Legal_Yearly" and isinstance(row[1], int):
            result.setdefault("legal_yearly", []).append(
                {"year": row[1], "settlement_usd": row[2], "defense_usd": row[3],
                 "fines_usd": row[4], "reg_defense_usd": row[5]})
    required = ("bi_incident_cost_usd", "bi_cost_usd", "bi_crisis_cost_usd",
                "bi_n_claims", "legal_settlement_usd", "legal_defense_usd",
                "regulatory_fines_usd", "regulatory_defense_usd", "legal_total_usd",
                "legal_n_claims", "recovery_expense_usd", "recovery_incident_cost_usd",
                "total_sme_claims",
                "recovery_n_claims", "crisis_services_pct",
                "bi_yearly", "recovery_yearly", "legal_yearly")
    missing = [k for k in required if k not in result]
    if missing:
        raise ValidationError(
            f"Severity_NetDiligence: incident cost context missing {missing} "
            "(workbook layout drift).")
    return result


def load_assumptions(wb):
    """Layer 4 Assumption Register. Category (Distribution choices /
    Independence & correlation assumptions / Scope limitations / Output &
    display definitions) is a workbook column, not a hardcoded mapping in
    code — same principle as Industry_Master's per-source columns."""
    ws = wb["Assumptions"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {name: i for i, name in enumerate(header) if name is not None}
    return [
        {"parameter": r[idx["Parameter"]], "value": r[idx["Value"]],
         "unit": r[idx["Unit"]], "reason": r[idx["Reason"]],
         "evidence": r[idx["Evidence"]], "category": r[idx["Category"]]}
        for r in rows[1:] if r[0] is not None
    ]


def load_evidence_register(wb):
    """Layer 4 Evidence Register: the 8 sources with SQR rating and reason."""
    ws = wb["Evidence_Register"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {name: i for i, name in enumerate(header) if name is not None}
    return [
        {"id": r[idx["Evidence_ID"]], "dataset": r[idx["Dataset"]],
         "source": r[idx["Source"]], "version": r[idx["Version"]],
         "sqr_rating": r[idx["SQR_Rating"]], "sqr_reason": r[idx["SQR_Reason"]]}
        for r in rows[1:] if r[0] is not None
    ]


def load_dbir_frequencies(wb):
    """Frequency_DBIR: DBIR industry -> [(incident_type, share), ...]."""
    ws = wb["Frequency_DBIR"]
    rows = list(ws.iter_rows(values_only=True))
    data = {}
    for r in rows[1:]:
        ind, incident, share = r[0], r[1], r[2]
        if ind is None:
            continue
        data.setdefault(ind, []).append((incident, float(share)))
    return data


_DATA = None


def _load_data():
    global _DATA
    if _DATA is None:
        wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
        baseline, ind_mult, rev_mult = load_frequency_tables(wb)
        ind_sev, rev_sev = load_severity_tables(wb)
        _DATA = {
            "baseline": baseline,
            "ind_mult": ind_mult,
            "rev_mult": rev_mult,
            "ind_sev": ind_sev,
            "rev_sev": rev_sev,
            "ssc_mult": load_ssc_multipliers(wb),
            "industry_master": load_industry_master(wb),
            "dbir": load_dbir_frequencies(wb),
            "rev_sev_t1": load_revenue_severity_t1(wb),
            "cost_components": load_cost_component_breakdown(wb),
            "assumptions": load_assumptions(wb),
            "evidence_register": load_evidence_register(wb),
            "cause_of_loss": load_cause_of_loss(wb),
            "incident_cost_context": load_incident_cost_context(wb),
        }
    return _DATA


def get_methodology():
    """Layer 4 read-only enrichment — NOT part of the Monte Carlo engine.

    Returns the full Assumption Register, Evidence Register, the
    Industry_Master confidence tally, the all-SME Top Causes of Loss
    ranking, and the Incident Cost context stats (BI/Legal/Recovery/Crisis
    Services share) — all independent of any specific calculation.
    """
    data = _load_data()
    tally = {"High": 0, "Medium": 0, "Low": 0, "None": 0}
    for row in data["industry_master"]:
        tally[row["mapping_confidence"]] = tally.get(row["mapping_confidence"], 0) + 1
    return {
        "assumptions": data["assumptions"],
        "evidence_register": data["evidence_register"],
        "industry_confidence_tally": tally,
        "industries": [
            {"industry_id": r["industry_id"], "display_name": r["display_name"],
             "mapping_confidence": r["mapping_confidence"]}
            for r in data["industry_master"]
        ],
        "cause_of_loss": data["cause_of_loss"],
        "incident_cost_context": data["incident_cost_context"],
    }


def build_loss_breakdown(master_row, freq_name, sev_name, components):
    """Layer 2 Loss Breakdown — independent of the Monte Carlo calculation.

    Sums Table B's Avg cost by Primary_Secondary tag (Mixed/Unclassified is
    reported separately, never folded into Primary or Secondary, per
    PROJECT_BRIEF: "Other Crisis Services" tag is Mixed/Unclassified — leave
    it out of both totals rather than guessing which bucket it belongs to).
    Only available when Applicable_Industry (a workbook column, not a
    hardcoded name) matches the selected industry under ANY of its three
    known names (display / IRIS-frequency / IRIS-severity), mirroring
    resolve_industry's own matching rule.
    """
    # Table B now holds rows for multiple industries (Manufacturing from the
    # specialized report, plus 11 more from Cyber Claims Study 2025 Table 6)
    # — filter to the rows whose own Applicable_Industry matches the
    # selected industry, rather than assuming the whole table is one
    # industry's data (that assumption broke as soon as a second industry
    # was added: every industry would otherwise see whichever industry's
    # rows happened to come first in the sheet).
    names = {str(master_row["display_name"]).lower(), str(freq_name).lower(),
             str(sev_name).lower()}
    matched = [c for c in components
               if str(c["Applicable_Industry"]).lower() in names]
    if not matched:
        available = sorted({c["Applicable_Industry"] for c in components})
        return {
            "available": False,
            "reason": f"Cost Component Breakdown is not populated for "
                      f"{master_row['display_name']} in Severity_NetDiligence "
                      f"— currently available for: {', '.join(available)}. "
                      "Showing this breakdown for other industries would "
                      "require fabricating a split that the source data "
                      "does not support.",
        }
    applicable = matched[0]["Applicable_Industry"]
    primary = [c for c in matched if c["Primary_Secondary"] == "Primary"]
    secondary = [c for c in matched if c["Primary_Secondary"] == "Secondary"]
    unclassified = [c for c in matched
                    if c["Primary_Secondary"] == "Mixed/Unclassified"]
    return {
        "available": True,
        "applicable_industry": applicable,
        "components": [
            {"name": c["Component"], "primary_secondary": c["Primary_Secondary"],
             "avg_usd": c["Avg"], "min_usd": c["Min"], "max_usd": c["Max"]}
            for c in matched
        ],
        "primary_total_usd": sum(c["Avg"] for c in primary),
        "secondary_total_usd": sum(c["Avg"] for c in secondary),
        "unclassified_total_usd": sum(c["Avg"] for c in unclassified),
        "source": f"{matched[0]['Source_Report']}, "
                  f"{matched[0]['Source_Section']}",
        "relationship_to_calculation": "independent_reference",
        "note": "This Primary/Secondary cost-component split comes from a "
                "separate NetDiligence per-incident study for this industry "
                "only. It is NOT an input to the Monte Carlo calculation "
                "above (which uses IRIS Frequency/Severity data) — it is "
                "shown as an independent reference pattern of how a typical "
                "incident's cost tends to break down, not a decomposition "
                "of the dollar figures calculated on this page.",
    }


def get_top_threat(industry):
    """Read-only presentation enrichment — NOT part of the Monte Carlo engine.

    Returns the largest DBIR attack-pattern share for the industry, resolved
    via Industry_Master's existing DBIR_Industry column (no mapping invented
    in code). If Industry_Master maps the industry to no DBIR category
    (DBIR_Industry=N/A: Agriculture, Management, Trade), returns
    available=False with the reason — no Overall/other fallback substituted,
    per product decision (Top Threat card shows "no data" for these).
    """
    data = _load_data()
    row = resolve_industry(industry, data["industry_master"])
    dbir_name = row["dbir_industry"]
    if dbir_name in (None, "N/A"):
        return {
            "available": False,
            "industry": row["display_name"],
            "dbir_industry": None,
            "reason": "Industry_Master maps this industry to no DBIR "
                      "category (DBIR_Industry=N/A) — DBIR 2026 does not "
                      "report it separately.",
        }
    patterns = data["dbir"].get(dbir_name)
    if not patterns:
        # Industry_Master names a DBIR category that Frequency_DBIR lacks:
        # workbook drift — surface it, do not guess.
        raise ValidationError(
            f"Industry_Master maps '{row['display_name']}' to DBIR category "
            f"'{dbir_name}', but Frequency_DBIR has no rows for it.")
    ranked = sorted(patterns, key=lambda t: -t[1])
    incident, share = ranked[0]
    return {
        "available": True,
        "industry": row["display_name"],
        "dbir_industry": dbir_name,
        "incident_type": incident,
        "share": share,
        # Full per-industry ranking (Frequency_DBIR has ~8 action patterns
        # per DBIR industry) — powers the Layer 2 "Top Causes of Loss" list
        # with a genuinely industry-specific breakdown. Note this is a
        # SHARE-OF-BREACHES metric (can sum >100%, DBIR incidents can match
        # multiple patterns), not the cost-share metric NetDiligence's
        # Cause_of_Loss table reports — the two are not interchangeable.
        "patterns": [{"incident_type": n, "share": s} for n, s in ranked],
        "source": "Verizon DBIR 2026 (Frequency_DBIR)",
    }


def resolve_industry(industry, master):
    """Match user input against Industry_Master display or IRIS names."""
    key = industry.strip().lower()
    for row in master:
        names = {str(row["display_name"]).lower(),
                 str(row["iris_frequency_name"]).lower(),
                 str(row["iris_severity_name"]).lower()}
        if key in names:
            if row["iris_frequency_name"] in (None, "N/A"):
                raise ValidationError(
                    f"Industry '{row['display_name']}' has no IRIS anchor "
                    f"(orphan industry, Mapping_Confidence="
                    f"{row['mapping_confidence']}) — no frequency baseline "
                    f"exists for it, so a calculation cannot be run.")
            return row
    available = sorted(r["display_name"] for r in master
                       if r["iris_frequency_name"] not in (None, "N/A"))
    raise ValidationError(
        f"Unknown industry '{industry}'. Available (Industry_Master display "
        f"names): {available}")


def validate_inputs(industry, revenue_class, ssc_grade, data):
    """PROJECT_BRIEF Step 0: Industry + Revenue required, SSC optional."""
    if not industry or not str(industry).strip():
        raise ValidationError(
            "Industry is required (Step 0 validation) — Monte Carlo is not "
            "run with partial input.")
    if not revenue_class or not str(revenue_class).strip():
        raise ValidationError(
            "Revenue tier is required (Step 0 validation) — Monte Carlo is "
            "not run with partial input.")
    row = resolve_industry(industry, data["industry_master"])
    if revenue_class not in data["rev_mult"]:
        raise ValidationError(
            f"Unknown revenue tier '{revenue_class}'. Must match a "
            f"Frequency_IRIS Table C label: {sorted(data['rev_mult'])}")
    if ssc_grade is not None and ssc_grade not in data["ssc_mult"]:
        raise ValidationError(
            f"Unknown SSC grade '{ssc_grade}'. Available (Frequency_SSC): "
            f"{sorted(data['ssc_mult'])}")
    return row


_CONF_LEVELS = ["Low", "Medium", "High"]


def derive_confidence(approx, mapping_confidence, ssc_applied):
    reasons = []
    if not approx and mapping_confidence == "High":
        level = "High"
        reasons.append("Published IRIS frequency baseline (one of 8 charted "
                       "industries) and Industry_Master Mapping_Confidence=High.")
    elif (not approx and mapping_confidence == "Medium") or \
         (approx and mapping_confidence == "High"):
        level = "Medium"
        if approx:
            reasons.append("Frequency baseline approximated (Overall x "
                           "Industry multiplier) despite High mapping confidence.")
        else:
            reasons.append("Published IRIS baseline, but Industry_Master "
                           "Mapping_Confidence=Medium.")
    else:
        level = "Low"
        reasons.append(f"Mapping_Confidence={mapping_confidence}"
                       f"{', approximated frequency baseline' if approx else ''}.")
    if ssc_applied:
        downgraded = _CONF_LEVELS[max(0, _CONF_LEVELS.index(level) - 1)]
        reasons.append("Downgraded one level: SSC multiplier applied, and "
                       "Frequency_SSC's source (EV_05) is single-source with "
                       "SQR=Low in the Evidence_Register.")
        level = downgraded
    return level, reasons


def calculate(industry, revenue_class, ssc_grade=None, scenario=None,
              n_runs=100_000, seed=42):
    """Run the full CRQ calculation for one Industry/Revenue/SSC combination.

    Returns a JSON-serializable dict; raises ValidationError on bad input.
    """
    data = _load_data()
    master_row = validate_inputs(industry, revenue_class, ssc_grade, data)
    freq_name = master_row["iris_frequency_name"]
    sev_name = master_row["iris_severity_name"]

    lef, approx, ssc_m = get_frequency(
        freq_name, revenue_class, data["baseline"], data["ind_mult"],
        data["rev_mult"], ssc_mult_table=data["ssc_mult"], ssc_grade=ssc_grade)
    try:
        mu, sigma, rev_note = get_severity_params(
            sev_name, revenue_class, data["ind_sev"], data["rev_sev"])
    except ValueError as e:
        # A per-sheet name missing from its sheet means Industry_Master and
        # the data sheets have drifted — surface the workbook's own naming,
        # do not guess a mapping in code.
        raise ValidationError(str(e))

    if scenario is not None and scenario not in SCENARIO_COVERAGE:
        raise ValidationError(
            f"Unknown scenario '{scenario}'. Available: "
            f"{sorted(SCENARIO_COVERAGE)}")

    # Design A revenue severity adjustment (Assumptions:
    # Severity_Revenue_Adjustment): shift the lognormal location by the
    # Table 1 median multiplier; sigma stays industry-based.
    t1_table, tier_map = data["rev_sev_t1"]
    m_rev, t1_tier, anchor_gm = revenue_severity_multiplier(
        revenue_class, t1_table, tier_map)
    mu_industry = mu
    mu = mu_industry + math.log(m_rev)

    losses = run_monte_carlo(lef, mu, sigma, n_runs=n_runs, seed=seed)
    stats = summarize(losses)

    # Loss Distribution chart data (Layer 2): bin the SAME simulated losses
    # array on a log10 axis (cyber loss is heavy-tailed; a linear-scale
    # histogram would be a single spike at/near zero and nothing else
    # visible). prob_zero_loss_year already covers the zero-loss mass
    # separately. No new calculation — this only bins the existing
    # `losses` array from run_monte_carlo().
    #
    # Bin edges are a FIXED $1-$10B log10 range (not each calculation's own
    # min/max) so the chart's SHAPE is comparable across calculations — a
    # low-sigma industry visibly clusters into a narrow spike and a
    # high-sigma one visibly spreads across more of the axis, instead of
    # every distribution being auto-stretched to fill the same 24 bins
    # regardless of its actual spread (that stretching is why the chart
    # used to look "the same shape" for every industry/revenue combo).
    # Values outside $1-$10B (rare — e.g. very high-sigma industry stacked
    # with the largest revenue tier and an SSC F grade) fold into the first/
    # last bin as an honest overflow rather than being clipped from view.
    positive = losses[losses > 0]
    histogram = None
    if positive.size > 0:
        log_vals = np.log10(positive)
        n_bins = 24
        FIXED_LO, FIXED_HI = 0.0, 10.0  # log10($1) .. log10($10B)
        clipped = np.clip(log_vals, FIXED_LO, FIXED_HI)
        edges = np.linspace(FIXED_LO, FIXED_HI, n_bins + 1)
        counts, edges = np.histogram(clipped, bins=edges)
        histogram = {
            "scale": "log10_usd",
            "bin_edges_usd": [float(10 ** e) for e in edges],
            "counts": [int(c) for c in counts],
            "below_range": int((log_vals < FIXED_LO).sum()),
            "above_range": int((log_vals > FIXED_HI).sum()),
        }

    # Loss Exceedance Curve (Layer 2 hero chart): P(annual loss > x) across a
    # fixed $1K-$10B log grid, taken straight from the SAME simulated `losses`
    # array (no new calculation - same "only reshapes the existing array"
    # rationale as `histogram` above). An LEC is the CRQ-standard presentation
    # and, unlike the histogram, needs no separate zero-loss spike: most years
    # are $0 for low-frequency industries, and that shows up simply as the
    # curve starting well below 100% (its left-edge value is P(loss > $1K) ~=
    # the fraction of years with a material loss; the gap up to 1.0 is the
    # quiet-year mass). Fixed range keeps the curve shape comparable across
    # calculations.
    lec = None
    if losses.size:
        LEC_LO, LEC_HI, LEC_N = 3.0, 10.0, 121  # log10($1K) .. log10($10B)
        xs = np.logspace(LEC_LO, LEC_HI, LEC_N)
        srt = np.sort(losses)
        # P(loss > x) = 1 - (count of losses <= x) / n
        ys = 1.0 - np.searchsorted(srt, xs, side="right") / float(losses.size)
        lec = {
            "loss_usd": [float(x) for x in xs],
            "exceed_prob": [float(y) for y in ys],
            "p_loss_gt_floor": float(ys[0]),
            "floor_usd": float(xs[0]),
            "max_usd": float(losses.max()),
        }

    # Per-incident severity percentiles (lognormal closed form). The P95 is
    # the Maximum Probable Loss / Recommended Coverage basis (Assumptions:
    # Recommended_Coverage_Basis) — equals industry P95 x revenue multiplier,
    # replacing the annual-aggregate P95 which is $0 whenever LEF < ~0.051.
    Z_90, Z_99 = 1.2816, 2.3263
    per_incident = {
        "median": math.exp(mu),
        "p90": math.exp(mu + Z_90 * sigma),
        "p95": math.exp(mu + Z_95 * sigma),
        "p99": math.exp(mu + Z_99 * sigma),
    }

    confidence, conf_reasons = derive_confidence(
        approx, master_row["mapping_confidence"], ssc_grade is not None)

    flags = []
    if approx:
        flags.append("Frequency baseline APPROXIMATED: no published IRIS "
                     "figure for this industry; using Overall (9.3%) x "
                     "Industry_Multiplier.")
    if master_row["mapping_confidence"] != "High":
        # Covers Low/Medium/None — not just Low/Medium — because mapping
        # confidence "None" (e.g. Management, Trade) drives derive_confidence()
        # to Low exactly the same way Low does; leaving it out would just
        # relocate the badge/count mismatch this flag exists to close.
        flags.append(
            "Industry_Master mapping for this industry is an unverified, "
            f"approximate cross-source mapping (Mapping_Confidence="
            f"{master_row['mapping_confidence']}) — label correspondence "
            "across IRIS/DBIR/IBM/NetDiligence has not been independently "
            "confirmed for this industry.")
    if ssc_grade is not None:
        flags.append(f"SSC adjustment (grade {ssc_grade}, x{ssc_m}) is "
                     "single-source (SQR=Low) — directional only.")
    flags.append(f"Revenue severity adjustment applied: median x{m_rev:.2f} "
                 f"(IRIS 2025 Table 1, tier '{t1_tier}', log-space "
                 "multiplier, GM anchor — marginal statistics, industry mix "
                 "not controlled).")
    flags.append("Coverage is Yes/No classification only — the specimen "
                 "policy has no deductible/sub-limit/aggregate amounts; no "
                 "net-loss dollar adjustment is modeled.")

    result = {
        "inputs": {
            "industry_id": master_row["industry_id"],
            "industry": master_row["display_name"],
            "iris_frequency_industry": freq_name,
            "iris_severity_industry": sev_name,
            "revenue_class": revenue_class,
            "ssc_grade": ssc_grade,
        },
        "frequency": {
            "lef": lef,
            "baseline": "approximated" if approx else "published_iris",
            "ssc_multiplier": ssc_m,
        },
        "severity": {
            "distribution": "Lognormal",
            "mu": mu,
            "sigma": sigma,
            "mu_industry": mu_industry,
            "revenue_adjustment": {
                "method": "log-space median multiplier (Design A)",
                "t1_tier": t1_tier,
                "t1_median_usd": t1_table[t1_tier][0],
                "t1_p95_usd": t1_table[t1_tier][1],
                "multiplier": m_rev,
                "anchor_gm_usd": anchor_gm,
            },
            "revenue_tier_typical_extreme_usd": list(rev_note) if rev_note else None,
        },
        "results": {
            "currency": "USD",
            "expected_annual_loss": float(stats["mean_ALE"]),
            "median": float(stats["median"]),
            "stddev": float(losses.std()),
            "p75": float(stats["p75"]),
            "p90": float(stats["p90"]),
            "p95": float(stats["p95"]),
            "p99": float(stats["p99"]),
            "prob_zero_loss_year": float(stats["prob_zero_loss_year"]),
            "simulation_runs": n_runs,
            "random_seed": seed,
            "per_incident_loss_usd": per_incident,
            "max_probable_loss_usd": per_incident["p95"],
            "recommended_coverage_usd": per_incident["p95"],
            "recommended_coverage_basis": "per-incident severity P95 "
                "(industry P95 x revenue multiplier), not annual-aggregate "
                "percentile",
            "histogram": histogram,
            "lec": lec,
        },
        "confidence": {
            "level": confidence,
            "reasons": conf_reasons,
            "mapping_confidence": master_row["mapping_confidence"],
        },
        "coverage": dict(SCENARIO_COVERAGE),
        "flags": flags,
        "loss_breakdown": build_loss_breakdown(
            master_row, freq_name, sev_name, data["cost_components"]),
    }
    if scenario is not None:
        result["scenario_check"] = {
            "scenario": scenario,
            "covered": SCENARIO_COVERAGE[scenario],
        }
    return result


def compact_view(result):
    """Minimal sample response shape requested for this step."""
    return {
        "industry": result["inputs"]["industry"],
        "expected_loss": round(result["results"]["expected_annual_loss"], 2),
        "confidence": result["confidence"]["level"],
        "coverage": result["coverage"],
    }


def main():
    parser = argparse.ArgumentParser(description="RiskSync CRQ JSON API (CLI)")
    parser.add_argument("--industry", default=None)
    parser.add_argument("--revenue", default=None)
    parser.add_argument("--ssc_grade", default=None)
    parser.add_argument("--scenario", default=None)
    parser.add_argument("--runs", type=int, default=100_000)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--compact", action="store_true",
                        help="Print only the minimal sample response shape")
    args = parser.parse_args()

    try:
        result = calculate(args.industry, args.revenue,
                           ssc_grade=args.ssc_grade, scenario=args.scenario,
                           n_runs=args.runs, seed=args.seed)
    except ValidationError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False, indent=2))
        sys.exit(1)

    out = compact_view(result) if args.compact else result
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
